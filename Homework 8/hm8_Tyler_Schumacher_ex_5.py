'''
Homework 8, Exercise 5
Tyler Schumacher
October 27, 2019
Write a program that converts an excel file to a csv file
'''
import os
import csv
import openpyxl

for excelFile in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object.
    if excelFile.endswith('.xlsx'):
        wb = openpyxl.load_workbook(excelFile)

        for sheetName in wb.get_sheet_names():
            # Loop through every sheet in the workbook.
            sheet = wb.get_sheet_by_name(sheetName)
            nameSplice = excelFile[: -5]

            # Create the CSV filename from the Excel filename and sheet title.
            csvFile = open(nameSplice + '_' + sheetName + '.csv', 'w', newline='')

            # Create the csv.writer object for this CSV file.
            csvWriter = csv.writer(csvFile)

            # Loop through every row in the sheet.
            for rowNum in range(1, sheet.max_row + 1):
                rowData = [] # append each cell to this list
                # Loop through each cell in the row.
                for colNum in range(1, sheet.max_column + 1):
                    # Append each cell's data to rowData.
                    rowData.append(sheet.cell(row=rowNum, column=colNum).value)

                # Write the rowData list to the CSV file.
                for row in rowData:
                    csvWriter.writerow(row)

            csvFile.close()